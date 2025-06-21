import pandas as pd
import numpy as np
import random
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class BeyannameSampling:
    """
    Beyanname örnekleme sınıfı. Gümrük beyannamelerinden
    belirli kriterlere göre örnekleme yaparak Excel raporlama yapar.
    """
    
    def __init__(self, df=None):
        """
        Args:
            df (pandas.DataFrame): Beyanname verileri içeren DataFrame
        """
        self.df = df
        self.unique_beyanname_df = None
        self.selected_beyannames = set()
        self.selection_reasons = {}  # Beyanname no -> seçim nedenleri listesi
        self.sampling_stats = {}  # İstatistikler
    
    def set_dataframe(self, df):
        """DataFrame'i ayarlar ve örnekleme için hazırlar"""
        self.df = df
        self._prepare_unique_beyannames()
        self.selected_beyannames = set()
        self.selection_reasons = {}
        self.sampling_stats = {}
    
    def _prepare_unique_beyannames(self):
        """
        Benzersiz beyanname numaralarına göre bir DataFrame hazırlar.
        Her beyanname için bir satır oluşturur.
        """
        if self.df is None or 'Beyanname_no' not in self.df.columns:
            return None
        
        # Beyanname numarasına göre grupla ve ilk satırı al
        self.unique_beyanname_df = self.df.drop_duplicates(subset=['Beyanname_no'])
        
        # Toplam beyanname sayısını kaydet
        self.sampling_stats['total_beyannames'] = len(self.unique_beyanname_df)
        
        return self.unique_beyanname_df
    
    def _add_selection_reason(self, beyanname_no, reason):
        """
        Seçilen beyannameye seçim nedeni ekler
        
        Args:
            beyanname_no (str): Beyanname numarası
            reason (str): Seçim nedeni
        """
        if beyanname_no not in self.selection_reasons:
            self.selection_reasons[beyanname_no] = []
        
        if reason not in self.selection_reasons[beyanname_no]:
            self.selection_reasons[beyanname_no].append(reason)
        
        # Seçilen beyannameler setine ekle
        self.selected_beyannames.add(beyanname_no)
    
    def run_sampling(self, min_sample_count=100, max_sample_count=150, sample_percentage=0.05):
        """
        Tüm kriterlere göre örnekleme yapar
        
        Args:
            min_sample_count (int): Minimum örnekleme sayısı
            max_sample_count (int): Maksimum örnekleme sayısı
            sample_percentage (float): Örnekleme yüzdesi (0-1 arası)
        
        Returns:
            pandas.DataFrame: Seçilen beyannameleri ve seçim nedenlerini içeren DataFrame
        """
        if self.df is None or self.unique_beyanname_df is None:
            raise ValueError("Veri yüklenmemiş veya benzersiz beyannameler oluşturulmamış")
        
        # Hedef örnekleme sayısını belirle
        total_count = len(self.unique_beyanname_df)
        target_sample_count = int(total_count * sample_percentage)
        target_sample_count = max(min_sample_count, min(max_sample_count, target_sample_count))
        
        self.sampling_stats['target_sample_count'] = target_sample_count
        
        # Temel seçim kriterlerini uygula
        self._sample_by_rejim_code()
        self._sample_by_sender()
        self._sample_by_gtip_code()
        self._sample_by_country()
        self._sample_by_highest_weight()
        self._sample_by_highest_value()
        self._sample_by_exemption_code()
        self._sample_by_simplified_procedure()
        self._sample_by_origin_code()
        self._sample_by_transport_type()
        
        # Özel kriterleri uygula
        self._sample_by_delivery_type()
        self._sample_by_foreign_expense()
        self._sample_by_exceptional_value()
        self._sample_by_onboard_process()
        self._sample_by_payment_method()
        self._sample_by_discount()
        self._sample_by_origin_proof_document()
        self._sample_by_atr_and_supplier_declaration()
        self._sample_by_set_classification()
        self._sample_by_supplier_origin_declaration()
        self._sample_by_special_purpose_exemption()
        self._sample_by_inward_outward_processing()
        
        # Hedef sayıya ulaşana kadar rastgele ek beyannameler seç
        self._random_sampling_to_target(target_sample_count)
        
        # Sonuçları DataFrame olarak hazırla
        return self._prepare_results_dataframe()
    
    def _sample_by_rejim_code(self):
        """Tüm rejim kodlarından en az %5 olacak şekilde örnekleme yapar"""
        if 'Rejim' not in self.df.columns:
            return
        
        # Rejim kodlarını say
        rejim_counts = self.df.groupby('Rejim')['Beyanname_no'].nunique().reset_index()
        rejim_counts.columns = ['Rejim', 'Beyanname_Sayisi']
        
        # Her rejim kodu için örnekleme yap
        for _, row in rejim_counts.iterrows():
            rejim_code = row['Rejim']
            beyanname_count = row['Beyanname_Sayisi']
            
            # Rejim koduna göre %5 örnekleme (minimum 1)
            sample_count = max(1, int(beyanname_count * 0.05))
            
            # Bu rejim kodlu benzersiz beyannameleri bul
            beyannames = self.df[self.df['Rejim'] == rejim_code]['Beyanname_no'].unique()
            
            # Mümkünse farklı firmalardan ve yüksek kıymetli beyannameleri önceliklendir
            if 'Adi_unvani' in self.df.columns and 'Fatura_miktari' in self.df.columns:
                # Firma ve kıymete göre grupla
                rejim_df = self.df[self.df['Rejim'] == rejim_code]
                firm_value_df = rejim_df.groupby(['Beyanname_no', 'Adi_unvani'])['Fatura_miktari'].sum().reset_index()
                
                # Kıymete göre sırala (yüksekten düşüğe)
                firm_value_df = firm_value_df.sort_values('Fatura_miktari', ascending=False)
                
                # Farklı firmaları önceliklendir
                selected_firms = set()
                selected_beyannames_for_rejim = []
                
                for _, brow in firm_value_df.iterrows():
                    beyanname_no = brow['Beyanname_no']
                    firm = brow['Adi_unvani']
                    
                    # Zaten yeterince beyanname seçtiyse çık
                    if len(selected_beyannames_for_rejim) >= sample_count:
                        break
                    
                    # Bu firma henüz seçilmemişse öncelikle seç
                    if firm not in selected_firms:
                        selected_firms.add(firm)
                        selected_beyannames_for_rejim.append(beyanname_no)
                        self._add_selection_reason(beyanname_no, f"Rejim Kodu: {rejim_code} - Farklı firma örneklemesi")
                
                # Hala yeterli sayıda değilse kalan kısmı yüksek kıymetlilerden tamamla
                if len(selected_beyannames_for_rejim) < sample_count:
                    remaining_beyannames = [b for b in beyannames if b not in selected_beyannames_for_rejim]
                    remaining_needed = sample_count - len(selected_beyannames_for_rejim)
                    
                    if remaining_needed > 0 and remaining_beyannames:
                        # Kalan beyannamelerden yüksek kıymetli olanları seç
                        remaining_df = rejim_df[rejim_df['Beyanname_no'].isin(remaining_beyannames)]
                        value_df = remaining_df.groupby('Beyanname_no')['Fatura_miktari'].sum().reset_index()
                        value_df = value_df.sort_values('Fatura_miktari', ascending=False)
                        
                        for idx, vrow in value_df.head(remaining_needed).iterrows():
                            beyanname_no = vrow['Beyanname_no']
                            self._add_selection_reason(beyanname_no, f"Rejim Kodu: {rejim_code} - Yüksek kıymet örneklemesi")
            else:
                # Firma ve kıymet bilgisi yoksa rastgele seç
                sample_size = min(sample_count, len(beyannames))
                if sample_size > 0:
                    sampled_beyannames = random.sample(list(beyannames), sample_size)
                    for beyanname_no in sampled_beyannames:
                        self._add_selection_reason(beyanname_no, f"Rejim Kodu: {rejim_code} - Rastgele örnekleme")
    
    def _sample_by_sender(self):
        """Tüm göndericilerden en az 1 adet beyanname seçer"""
        sender_columns = ["Gonderen", "Gonderen_adi", "Gonderen_firma", "Adi_unvani", "Ihracatci"]
        
        # Mevcut gönderici sütununu bul
        sender_column = None
        for col in sender_columns:
            if col in self.df.columns:
                sender_column = col
                break
        
        if not sender_column:
            return
        
        # Gönderici firmaları benzersiz olarak al
        senders = self.df[sender_column].dropna().unique()
        
        for sender in senders:
            # Bu göndericiye ait beyannameler
            sender_beyannames = self.df[self.df[sender_column] == sender]['Beyanname_no'].unique()
            
            # Zaten seçilmiş bir beyanname var mı kontrol et
            already_selected = False
            for beyanname_no in sender_beyannames:
                if beyanname_no in self.selected_beyannames:
                    already_selected = True
                    break
            
            # Henüz bu göndericiden beyanname seçilmemişse, bir tane seç
            if not already_selected and len(sender_beyannames) > 0:
                selected_beyanname = random.choice(sender_beyannames)
                self._add_selection_reason(selected_beyanname, f"Gönderici: {sender} - Her göndericiden en az bir beyanname kriteri")
    
    def _sample_by_gtip_code(self):
        """Tüm GTIP kodlarından benzersiz olacak şekilde en az bir GTIP kodu seçer"""
        if 'Gtip' not in self.df.columns:
            return
        
        # Benzersiz GTIP kodları
        gtip_codes = self.df['Gtip'].dropna().unique()
        
        for gtip in gtip_codes:
            # Bu GTIP koduna sahip beyannameler
            gtip_beyannames = self.df[self.df['Gtip'] == gtip]['Beyanname_no'].unique()
            
            # Zaten bu GTIP'li beyanname seçilmiş mi kontrol et
            already_selected = False
            for beyanname_no in gtip_beyannames:
                if beyanname_no in self.selected_beyannames:
                    already_selected = True
                    break
            
            # Henüz bu GTIP'den beyanname seçilmemişse, bir tane seç
            if not already_selected and len(gtip_beyannames) > 0:
                selected_beyanname = random.choice(gtip_beyannames)
                self._add_selection_reason(selected_beyanname, f"GTIP: {gtip} - Her GTIP kodundan en az bir beyanname kriteri")
    
    def _sample_by_country(self):
        """Farklı ülkelerden beyanname seçer"""
        country_columns = ["Mensei_ulke", "Cikis_ulkesi", "Ihracat_ulkesi"]
        
        # Mevcut ülke sütununu bul
        country_column = None
        for col in country_columns:
            if col in self.df.columns:
                country_column = col
                break
        
        if not country_column:
            return
        
        # Benzersiz ülkeleri al
        countries = self.df[country_column].dropna().unique()
        
        for country in countries:
            # Bu ülkeye ait beyannameler
            country_beyannames = self.df[self.df[country_column] == country]['Beyanname_no'].unique()
            
            # Zaten bu ülkeden beyanname seçilmiş mi kontrol et
            already_selected = False
            for beyanname_no in country_beyannames:
                if beyanname_no in self.selected_beyannames:
                    already_selected = True
                    break
            
            # Henüz bu ülkeden beyanname seçilmemişse, bir tane seç
            if not already_selected and len(country_beyannames) > 0:
                selected_beyanname = random.choice(country_beyannames)
                self._add_selection_reason(selected_beyanname, f"Ülke: {country} - Farklı ülkelerden beyanname kriteri")
    
    def _sample_by_highest_weight(self):
        """En yüksek ağırlıktaki 5 adet beyannameyi seçer"""
        if 'Brut_agirlik' not in self.df.columns:
            return
        
        # Beyanname başına toplam brüt ağırlık hesapla
        weight_df = self.df.groupby('Beyanname_no')['Brut_agirlik'].sum().reset_index()
        
        # Ağırlığa göre sırala ve en yüksek 5 adet beyanname seç
        weight_df = weight_df.sort_values('Brut_agirlik', ascending=False)
        
        for _, row in weight_df.head(5).iterrows():
            beyanname_no = row['Beyanname_no']
            weight = row['Brut_agirlik']
            self._add_selection_reason(beyanname_no, f"En yüksek ağırlık: {weight} - İlk 5 beyanname")
    
    def _sample_by_highest_value(self):
        """En yüksek kıymetteki 5 adet beyannameyi seçer"""
        if 'Fatura_miktari' not in self.df.columns:
            return
        
        # Beyanname başına toplam kıymet hesapla
        value_df = self.df.groupby('Beyanname_no')['Fatura_miktari'].sum().reset_index()
        
        # Kıymete göre sırala ve en yüksek 5 adet beyanname seç
        value_df = value_df.sort_values('Fatura_miktari', ascending=False)
        
        for _, row in value_df.head(5).iterrows():
            beyanname_no = row['Beyanname_no']
            value = row['Fatura_miktari']
            self._add_selection_reason(beyanname_no, f"En yüksek kıymet: {value} - İlk 5 beyanname")
    
    def _sample_by_exemption_code(self):
        """Farklı muafiyet kodlarından beyanname seçer"""
        exemption_columns = ["Muafiyet_kodu", "Muafiyet", "Muafiyet1", "Muafiyet2"]
        
        # Mevcut muafiyet sütununu bul
        exemption_column = None
        for col in exemption_columns:
            if col in self.df.columns:
                exemption_column = col
                break
        
        if not exemption_column:
            return
        
        # Benzersiz muafiyet kodlarını al
        exemption_codes = self.df[exemption_column].dropna().unique()
        
        for exemption in exemption_codes:
            # Bu muafiyet koduna sahip beyannameler
            exemption_beyannames = self.df[self.df[exemption_column] == exemption]['Beyanname_no'].unique()
            
            # Zaten bu muafiyet kodlu beyanname seçilmiş mi kontrol et
            already_selected = False
            for beyanname_no in exemption_beyannames:
                if beyanname_no in self.selected_beyannames:
                    already_selected = True
                    break
            
            # Henüz bu muafiyet kodundan beyanname seçilmemişse, bir tane seç
            if not already_selected and len(exemption_beyannames) > 0:
                selected_beyanname = random.choice(exemption_beyannames)
                self._add_selection_reason(selected_beyanname, f"Muafiyet Kodu: {exemption} - Farklı muafiyet kodlarından beyanname kriteri")
    
    def _sample_by_simplified_procedure(self):
        """Tüm basitleştirilmiş usuldeki beyannamelerden en az bir adet seçer"""
        if 'Basitlestirilmis_usul' not in self.df.columns and 'Basitlestirilmis_usul_kodu' not in self.df.columns:
            return
        
        # Hangi sütunu kullanacağımızı belirle
        column = 'Basitlestirilmis_usul' if 'Basitlestirilmis_usul' in self.df.columns else 'Basitlestirilmis_usul_kodu'
        
        # Basitleştirilmiş usul değerlerini al
        simplified_codes = self.df[column].dropna().unique()
        
        for code in simplified_codes:
            # Bu usul koduna sahip beyannameler
            simplified_beyannames = self.df[self.df[column] == code]['Beyanname_no'].unique()
            
            # En az bir adet beyanname seç
            if len(simplified_beyannames) > 0:
                selected_beyanname = random.choice(simplified_beyannames)
                self._add_selection_reason(selected_beyanname, f"Basitleştirilmiş Usul: {code} - Her usulden en az bir beyanname kriteri")
    
    def _sample_by_origin_code(self):
        """Farklı menşe kodlarından beyanname seçer"""
        if 'Mensei_ulke' not in self.df.columns:
            return
        
        # Menşe kodlarını al
        origin_codes = self.df['Mensei_ulke'].dropna().unique()
        
        for origin in origin_codes:
            # Bu menşe koduna sahip beyannameler
            origin_beyannames = self.df[self.df['Mensei_ulke'] == origin]['Beyanname_no'].unique()
            
            # Zaten bu menşe kodlu beyanname seçilmiş mi kontrol et
            already_selected = False
            for beyanname_no in origin_beyannames:
                if beyanname_no in self.selected_beyannames:
                    already_selected = True
                    break
            
            # Henüz bu menşe kodundan beyanname seçilmemişse, bir tane seç
            if not already_selected and len(origin_beyannames) > 0:
                selected_beyanname = random.choice(origin_beyannames)
                self._add_selection_reason(selected_beyanname, f"Menşe Kodu: {origin} - Farklı menşe kodlarından beyanname kriteri")
    
    def _sample_by_transport_type(self):
        """Tüm taşıma türlerinden en az 1 adet beyanname seçer"""
        transport_columns = ["Tasima_sekli", "Tasima_araci", "Tasima_turu"]
        
        # Mevcut taşıma şekli sütununu bul
        transport_column = None
        for col in transport_columns:
            if col in self.df.columns:
                transport_column = col
                break
        
        if not transport_column:
            return
        
        # Taşıma türlerini al
        transport_types = self.df[transport_column].dropna().unique()
        
        for transport in transport_types:
            # Bu taşıma türüne sahip beyannameler
            transport_beyannames = self.df[self.df[transport_column] == transport]['Beyanname_no'].unique()
            
            # Zaten bu taşıma türünden beyanname seçilmiş mi kontrol et
            already_selected = False
            for beyanname_no in transport_beyannames:
                if beyanname_no in self.selected_beyannames:
                    already_selected = True
                    break
            
            # Henüz bu taşıma türünden beyanname seçilmemişse, bir tane seç
            if not already_selected and len(transport_beyannames) > 0:
                selected_beyanname = random.choice(transport_beyannames)
                self._add_selection_reason(selected_beyanname, f"Taşıma Türü: {transport} - Her taşıma türünden en az bir beyanname kriteri")
    
    def _sample_by_delivery_type(self):
        """Tüm teslim şekillerinden en az 1 adet beyanname seçer"""
        delivery_columns = ["Teslim_sekli"]
        
        # Mevcut teslim şekli sütununu bul
        delivery_column = None
        for col in delivery_columns:
            if col in self.df.columns:
                delivery_column = col
                break
        
        if not delivery_column:
            return
        
        # Teslim şekillerini al
        delivery_types = self.df[delivery_column].dropna().unique()
        
        # Sık kullanılan teslim şekillerini belirleme
        delivery_counts = self.df.groupby(delivery_column)['Beyanname_no'].nunique().reset_index()
        delivery_counts.columns = ['Teslim_Sekli', 'Beyanname_Sayisi']
        delivery_counts = delivery_counts.sort_values('Beyanname_Sayisi', ascending=False)
        
        # En çok kullanılan 3 teslim şeklinden 3'er adet örnek seç
        top_deliveries = delivery_counts.head(3)['Teslim_Sekli'].tolist()
        
        for delivery in delivery_types:
            # Bu teslim şekline sahip beyannameler
            delivery_beyannames = self.df[self.df[delivery_column] == delivery]['Beyanname_no'].unique()
            
            # Seçilecek beyanname sayısını belirle
            sample_count = 3 if delivery in top_deliveries else 1
            sample_size = min(sample_count, len(delivery_beyannames))
            
            if sample_size > 0:
                sampled_beyannames = random.sample(list(delivery_beyannames), sample_size)
                for beyanname_no in sampled_beyannames:
                    if delivery in top_deliveries:
                        self._add_selection_reason(beyanname_no, f"Teslim Şekli: {delivery} - En çok kullanılan teslim şekillerinden örnek")
                    else:
                        self._add_selection_reason(beyanname_no, f"Teslim Şekli: {delivery} - Her teslim şeklinden en az bir beyanname kriteri")
    
    def _sample_by_foreign_expense(self):
        """Yurt dışı gider bölümünde kıymet beyanı olanlardan örnekleme yapar"""
        foreign_expense_columns = ["Yurtdisi_gider", "Yurtdisi_gider_aciklama", "Royalti", "Lisans"]
        
        # Yurt dışı gider ve açıklama sütunlarını bul
        expense_column = None
        explanation_column = None
        
        for col in self.df.columns:
            if any(keyword in col.lower() for keyword in ["yurtdisi_gider", "royalti", "lisans"]):
                if "aciklama" in col.lower() or "açıklama" in col.lower():
                    explanation_column = col
                else:
                    expense_column = col
        
        # Eğer uygun sütunlar bulunamazsa çık
        if not expense_column and not explanation_column:
            return
        
        # Yurt dışı gider olan beyannameleri bul
        expense_beyannames = []
        
        # Gider sütunu varsa ve sayısalsa, sıfırdan büyük değerleri kontrol et
        if expense_column and pd.api.types.is_numeric_dtype(self.df[expense_column]):
            expense_beyannames.extend(self.df[self.df[expense_column] > 0]['Beyanname_no'].unique())
        
        # Açıklama sütunu varsa, "royalti", "lisans" içerenleri kontrol et
        if explanation_column:
            keywords = ["royalti", "lisans", "license", "royalty", "know-how", "franchise"]
            for keyword in keywords:
                matching_beyannames = self.df[self.df[explanation_column].str.lower().str.contains(keyword, na=False)]['Beyanname_no'].unique()
                expense_beyannames.extend(matching_beyannames)
        
        # Benzersiz beyannameleri al
        expense_beyannames = list(set(expense_beyannames))
        
        # En az 5 beyanname seç
        sample_size = min(5, len(expense_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(expense_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "Yurt dışı gider/royalti/lisans ödemesi olan beyanname")
    
    def _sample_by_exceptional_value(self):
        """İstisnai kıymetle beyan olan beyannameleri seçer"""
        # 44 numaralı hanenin bulunduğu sütunları belirleme
        hane44_columns = []
        for col in self.df.columns:
            if any(keyword in col.lower() for keyword in ["belge", "dokuman", "aciklama", "44"]):
                hane44_columns.append(col)
        
        if not hane44_columns:
            return
        
        # İstisnai kıymet ile ilgili anahtar kelimeler
        keywords = ["istisnai kıymet", "istisnai", "kiymet istisnasi", "kıymet istisnası"]
        
        exceptional_beyannames = []
        
        # Her bir potansiyel sütunu kontrol et
        for col in hane44_columns:
            if pd.api.types.is_string_dtype(self.df[col]):
                for keyword in keywords:
                    matching_beyannames = self.df[self.df[col].str.lower().str.contains(keyword, na=False)]['Beyanname_no'].unique()
                    exceptional_beyannames.extend(matching_beyannames)
        
        # Benzersiz beyannameleri al
        exceptional_beyannames = list(set(exceptional_beyannames))
        
        # En az 5 beyanname seç
        sample_size = min(5, len(exceptional_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(exceptional_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "İstisnai kıymet ile beyan edilen beyanname")
    
    def _sample_by_onboard_process(self):
        """Taşıt üstü işlem yapılan beyannameleri seçer (Basitleştirilmiş işlem kodu 3 olanlar)"""
        simplified_columns = ["Basitlestirilmis_usul", "Basitlestirilmis_usul_kodu", "Islem_kodu"]
        
        # Uygun sütunu belirle
        simplified_column = None
        for col in simplified_columns:
            if col in self.df.columns:
                simplified_column = col
                break
        
        if not simplified_column:
            return
        
        # Kod 3 olan beyannameleri bul
        onboard_beyannames = self.df[self.df[simplified_column] == '3']['Beyanname_no'].unique()
        
        # En az 5 beyanname seç
        sample_size = min(5, len(onboard_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(list(onboard_beyannames), sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "Taşıt üstü işlem yapılan beyanname (Basitleştirilmiş işlem kodu 3)")
    
    def _sample_by_payment_method(self):
        """Tüm ödeme şekillerinden en az 1 adet beyanname seçer"""
        payment_columns = ["Odeme", "Odeme_sekli", "Odeme_yontemi"]
        
        # Mevcut ödeme şekli sütununu bul
        payment_column = None
        for col in payment_columns:
            if col in self.df.columns:
                payment_column = col
                break
        
        if not payment_column:
            return
        
        # Ödeme şekillerini al
        payment_methods = self.df[payment_column].dropna().unique()
        
        # Sık kullanılan ödeme şekillerini belirleme
        payment_counts = self.df.groupby(payment_column)['Beyanname_no'].nunique().reset_index()
        payment_counts.columns = ['Odeme_Sekli', 'Beyanname_Sayisi']
        payment_counts = payment_counts.sort_values('Beyanname_Sayisi', ascending=False)
        
        # En çok kullanılan ödeme şeklini belirleyelim
        most_common_payment = None
        if not payment_counts.empty:
            most_common_payment = payment_counts.iloc[0]['Odeme_Sekli']
        
        for payment in payment_methods:
            # Bu ödeme şekline sahip beyannameler
            payment_beyannames = self.df[self.df[payment_column] == payment]['Beyanname_no'].unique()
            
            # Her ödeme şeklinden en az 1 beyanname seç
            # En yaygın ödeme şeklinden 3 beyanname seç
            sample_size = min(3 if payment == most_common_payment else 1, len(payment_beyannames))
            
            if sample_size > 0:
                sampled_beyannames = random.sample(list(payment_beyannames), sample_size)
                for beyanname_no in sampled_beyannames:
                    if payment == most_common_payment:
                        self._add_selection_reason(beyanname_no, f"Ödeme Şekli: {payment} - En yaygın ödeme şekli")
                    else:
                        self._add_selection_reason(beyanname_no, f"Ödeme Şekli: {payment} - Her ödeme şeklinden en az bir beyanname kriteri")
    
    def _sample_by_discount(self):
        """İskonto yapılan beyannameleri seçer (yurt dışı gider sütununda "iskonto" yazanlar)"""
        # Yurt dışı gider ve açıklama sütunlarını bul
        expense_columns = []
        for col in self.df.columns:
            if any(keyword in col.lower() for keyword in ["yurtdisi_gider", "gider", "iskonto", "indirim", "discount"]):
                expense_columns.append(col)
        
        if not expense_columns:
            return
        
        discount_beyannames = []
        
        # Her bir potansiyel sütunu kontrol et
        for col in expense_columns:
            if pd.api.types.is_string_dtype(self.df[col]):
                matching_beyannames = self.df[self.df[col].str.lower().str.contains("iskonto|indirim|discount", na=False, regex=True)]['Beyanname_no'].unique()
                discount_beyannames.extend(matching_beyannames)
        
        # Benzersiz beyannameleri al
        discount_beyannames = list(set(discount_beyannames))
        
        # En az 5 beyanname seç
        sample_size = min(5, len(discount_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(discount_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "İskonto yapılan beyanname")
    
    def _sample_by_origin_proof_document(self):
        """Menşe ispat belgesi kullanan beyannameleri seçer (0302, 0807, 0307 gibi belge kodları)"""
        # Doküman kod sütunlarını bul
        doc_code_columns = []
        for col in self.df.columns:
            if "dokuman" in col.lower() and "kod" in col.lower():
                doc_code_columns.append(col)
        
        if not doc_code_columns:
            return
        
        # Menşe ispat belge kodları
        origin_proof_codes = ['0302', '0807', '0307']
        
        origin_proof_beyannames = []
        
        # Her bir doküman kod sütununu kontrol et
        for col in doc_code_columns:
            for code in origin_proof_codes:
                matching_beyannames = self.df[self.df[col] == code]['Beyanname_no'].unique()
                origin_proof_beyannames.extend(matching_beyannames)
        
        # Benzersiz beyannameleri al
        origin_proof_beyannames = list(set(origin_proof_beyannames))
        
        # En az 1 beyanname seç (mümkünse farklı belge kodları için)
        if origin_proof_beyannames:
            selected_count = 0
            selected_codes = set()
            
            # Her menşe ispat belge kodu için en az 1 beyanname seç
            for code in origin_proof_codes:
                code_beyannames = []
                
                for col in doc_code_columns:
                    matching = self.df[self.df[col] == code]['Beyanname_no'].unique()
                    code_beyannames.extend(matching)
                
                code_beyannames = list(set(code_beyannames))
                
                if code_beyannames:
                    selected_beyanname = random.choice(code_beyannames)
                    self._add_selection_reason(selected_beyanname, f"Menşe ispat belgesi kullanan beyanname (Belge kodu: {code})")
                    selected_count += 1
                    selected_codes.add(code)
            
            # Eğer hiç seçilmediyse, genel havuzdan seç
            if selected_count == 0 and origin_proof_beyannames:
                selected_beyanname = random.choice(origin_proof_beyannames)
                self._add_selection_reason(selected_beyanname, "Menşe ispat belgesi kullanan beyanname")
    
    def _sample_by_atr_and_supplier_declaration(self):
        """A.TR ve Tedarikçi Beyanı olan beyannameleri seçer (0301 ve 0819 kodları)"""
        # Doküman kod sütunlarını bul
        doc_code_columns = []
        for col in self.df.columns:
            if "dokuman" in col.lower() and "kod" in col.lower():
                doc_code_columns.append(col)
        
        if not doc_code_columns:
            return
        
        # A.TR ve Tedarikçi Beyanı kodları
        atr_code = '0301'
        supplier_code = '0819'
        
        # Her iki belgeyi de içeren beyannameleri bul
        atr_supplier_beyannames = []
        
        for beyanname_no in self.df['Beyanname_no'].unique():
            beyanname_df = self.df[self.df['Beyanname_no'] == beyanname_no]
            
            has_atr = False
            has_supplier = False
            
            for col in doc_code_columns:
                if atr_code in beyanname_df[col].values:
                    has_atr = True
                if supplier_code in beyanname_df[col].values:
                    has_supplier = True
            
            if has_atr and has_supplier:
                atr_supplier_beyannames.append(beyanname_no)
        
        # En az 5 beyanname seç
        sample_size = min(5, len(atr_supplier_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(atr_supplier_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "A.TR Dolaşım Belgesi ve Tedarikçi Beyanı içeren beyanname")
    
    def _sample_by_set_classification(self):
        """Set halinde sınıflandırılan eşyaları seçer"""
        # Miktar birimi ve açıklama sütunlarını bul
        quantity_columns = ["Miktar_birimi", "Olcu_birimi"]
        description_columns = ["Aciklama", "Ticari_tanimi", "Esya_tanimi"]
        
        # Uygun sütunları belirleme
        quantity_column = None
        for col in quantity_columns:
            if col in self.df.columns:
                quantity_column = col
                break
        
        description_cols = []
        for col in description_columns:
            if col in self.df.columns:
                description_cols.append(col)
        
        set_beyannames = []
        
        # Miktar birimi "set" olanları kontrol et
        if quantity_column:
            set_quantity_beyannames = self.df[self.df[quantity_column].str.lower().str.contains("set", na=False)]['Beyanname_no'].unique()
            set_beyannames.extend(set_quantity_beyannames)
        
        # Açıklamalarda "set" içerenleri kontrol et
        for col in description_cols:
            if pd.api.types.is_string_dtype(self.df[col]):
                set_description_beyannames = self.df[self.df[col].str.lower().str.contains("set", na=False)]['Beyanname_no'].unique()
                set_beyannames.extend(set_description_beyannames)
        
        # Benzersiz beyannameleri al
        set_beyannames = list(set(set_beyannames))
        
        # En az 5 beyanname seç
        sample_size = min(5, len(set_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(set_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "Set halinde sınıflandırılan eşya")
    
    def _sample_by_supplier_origin_declaration(self):
        """Tedarikçi beyanı/menşe beyanı olan beyannameleri seçer (0876 veya 0842 kodları)"""
        # Doküman kod sütunlarını bul
        doc_code_columns = []
        for col in self.df.columns:
            if "dokuman" in col.lower() and "kod" in col.lower():
                doc_code_columns.append(col)
        
        if not doc_code_columns:
            return
        
        # Tedarikçi beyanı ve menşe beyanı kodları
        declaration_codes = ['0876', '0842']
        
        declaration_beyannames = []
        
        # Her bir doküman kod sütununu kontrol et
        for col in doc_code_columns:
            for code in declaration_codes:
                matching_beyannames = self.df[self.df[col] == code]['Beyanname_no'].unique()
                declaration_beyannames.extend(matching_beyannames)
        
        # Benzersiz beyannameleri al
        declaration_beyannames = list(set(declaration_beyannames))
        
        # En az 5 beyanname seç
        sample_size = min(5, len(declaration_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(declaration_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "Tedarikçi beyanı/menşe beyanı olan beyanname")
    
    def _sample_by_special_purpose_exemption(self):
        """Belirli amaç/nihai kullanım için indirimli veya sıfır vergi oranı (muafiyet "nkul", "nkul1", "nkul2" içerenler)"""
        exemption_columns = ["Muafiyet_kodu", "Muafiyet", "Muafiyet1", "Muafiyet2"]
        
        # Uygun sütunu belirle
        exemption_column = None
        for col in exemption_columns:
            if col in self.df.columns:
                exemption_column = col
                break
        
        if not exemption_column:
            return
        
        # "nkul" içeren beyannameleri bul
        nkul_beyannames = []
        if pd.api.types.is_string_dtype(self.df[exemption_column]):
            nkul_beyannames = self.df[self.df[exemption_column].str.lower().str.contains("nkul", na=False)]['Beyanname_no'].unique()
        
        # En az 5 beyanname seç
        sample_size = min(5, len(nkul_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(list(nkul_beyannames), sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "Belirli amaç/nihai kullanım için indirimli/sıfır vergi oranlı beyanname")
    
    def _sample_by_inward_outward_processing(self):
        """Dahilde veya Hariçte İşleme Rejimi beyannameleri (rejim kodu 5100, 5171, 2100)"""
        if 'Rejim' not in self.df.columns:
            return
        
        # Dahilde/Hariçte İşleme Rejimi kodları
        processing_codes = ['5100', '5171', '2100']
        
        processing_beyannames = []
        
        # Her bir rejim kodunu kontrol et
        for code in processing_codes:
            matching_beyannames = self.df[self.df['Rejim'] == code]['Beyanname_no'].unique()
            processing_beyannames.extend(matching_beyannames)
        
        # Benzersiz beyannameleri al
        processing_beyannames = list(set(processing_beyannames))
        
        # En az 5 beyanname seç
        sample_size = min(5, len(processing_beyannames))
        if sample_size > 0:
            sampled_beyannames = random.sample(processing_beyannames, sample_size)
            for beyanname_no in sampled_beyannames:
                self._add_selection_reason(beyanname_no, "Dahilde/Hariçte İşleme Rejimi beyannamesi")
    
    def _random_sampling_to_target(self, target_sample_count):
        """
        Hedef sayıya ulaşana kadar rastgele ek beyannameler seçer
        
        Args:
            target_sample_count (int): Hedef örnekleme sayısı
        """
        if self.df is None or self.unique_beyanname_df is None:
            raise ValueError("Veri yüklenmemiş veya benzersiz beyannameler oluşturulmamış")
        
        # Hedef sayıya ulaşana kadar rastgele ek beyannameler seç
        while len(self.selected_beyannames) < target_sample_count:
            remaining_beyannames = self.unique_beyanname_df[~self.unique_beyanname_df['Beyanname_no'].isin(self.selected_beyannames)]['Beyanname_no'].unique()
            if len(remaining_beyannames) > 0:
                selected_beyanname = random.choice(remaining_beyannames)
                self._add_selection_reason(selected_beyanname, "Rastgele örnekleme")
    
    def _prepare_results_dataframe(self):
        """
        Seçilen beyannameleri ve seçim nedenlerini DataFrame olarak hazırlar
        
        Returns:
            pandas.DataFrame: Seçilen beyannameleri ve seçim nedenlerini içeren DataFrame
        """
        if self.df is None or self.unique_beyanname_df is None:
            raise ValueError("Veri yüklenmemiş veya benzersiz beyannameler oluşturulmamış")
        
        # Seçilen beyannamelerin benzersiz DataFrame'ini hazırla
        selected_unique_df = self.unique_beyanname_df[self.unique_beyanname_df['Beyanname_no'].isin(self.selected_beyannames)]
        
        # Seçim nedenlerini DataFrame olarak hazırla
        reasons_df = pd.DataFrame({
            'Beyanname_no': list(self.selection_reasons.keys()),
            'Seçim_Nedenleri': [', '.join(reasons) for reasons in self.selection_reasons.values()]
        })
        
        # Seçilen beyannameler ve nedenlerini birleştir
        results_df = pd.merge(selected_unique_df, reasons_df, on='Beyanname_no', how='left')
        
        return results_df
    
    def export_to_excel(self, output_path=None):
        """
        Örnekleme sonuçlarını Excel dosyasına aktarır
        
        Args:
            output_path (str, optional): Excel dosyasının kaydedileceği yol. Belirtilmezse, geçerli tarih ile oluşturulur.
        
        Returns:
            str: Oluşturulan Excel dosyasının yolu
        """
        if not self.selected_beyannames:
            raise ValueError("Henüz örnekleme yapılmamış")
        
        # Çıktı yolu belirtilmemişse, geçerli tarih ile oluştur
        if output_path is None:
            current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"Beyanname_Ornekleme_{current_date}.xlsx"
        
        writer = None
        
        try:
            # Dosyanın yaratılabilir olduğunu kontrol et
            with open(output_path, 'a') as test_file:
                pass
            
            # Seçilen beyanname numaralarına göre benzersiz beyannameleri al
            # Güvenlik için kopyasını oluştur (orijinal veriler değişmesin)
            if self.unique_beyanname_df is None or len(self.unique_beyanname_df) == 0:
                raise ValueError("Beyanname verileri bulunamadı")
            
            # Hafıza optimizasyonu için sadece gereken sütunları seç
            # Tüm sütunlar yerine sadece ihtiyaç duyulan sütunlar işlenecek
            essential_columns = ['Beyanname_no']
            
            # Tarih sütununu bul
            date_column = None
            for col in self.unique_beyanname_df.columns:
                if 'tarih' in col.lower() or 'tescil' in col.lower():
                    date_column = col
                    essential_columns.append(col)
                    break
            
            # Diğer yararlı sütunları ekle
            useful_columns = ['Rejim', 'Gtip', 'Mensei_ulke', 'Odeme_sekli', 'Teslim_sekli', 
                            'Gonderen', 'Gonderen_adi', 'Alici', 'Alici_adi', 'Fatura_miktari', 
                            'Brut_agirlik', 'Net_agirlik']
            
            for col in useful_columns:
                if col in self.unique_beyanname_df.columns and col not in essential_columns:
                    essential_columns.append(col)
            
            # Optimize edilmiş veri seti oluştur
            needed_columns = [col for col in essential_columns if col in self.unique_beyanname_df.columns]
            if not needed_columns:
                needed_columns = ['Beyanname_no']  # En azından beyanname numarası kesin olmalı
            
            try:
                # Veri kopyasını al ve filtrele
                selected_df_copy = self.unique_beyanname_df[needed_columns].copy()
                selected_unique_df = selected_df_copy[selected_df_copy['Beyanname_no'].isin(self.selected_beyannames)]
                
                # Veriyi temizle
                selected_unique_df = selected_unique_df.fillna('')  # NaN değerleri temizle
                
                # Optimize etmek için kategorik sütunları dönüştür
                for col in selected_unique_df.columns:
                    if selected_unique_df[col].dtype == object:
                        try:
                            selected_unique_df[col] = selected_unique_df[col].astype('category')
                        except:
                            pass  # Dönüştürme başarısız olursa atla
            except Exception as e:
                print(f"Veri hazırlama hatası: {str(e)}")
                # Fallback: En azından beyanname numaralarını içeren DF oluştur
                selected_unique_df = pd.DataFrame({'Beyanname_no': list(self.selected_beyannames)})
            
            # Seçilen beyannameleri ve nedenlerini DataFrame olarak hazırla
            summary_data = []
            for beyanname_no in self.selected_beyannames:
                # Beyanname numarasına sahip kayıt var mı kontrol et
                matching_rows = selected_unique_df[selected_unique_df['Beyanname_no'] == beyanname_no]
                if matching_rows.empty:
                    continue
                
                # Sadece gerekli bilgileri al
                summary_row = {
                    'Beyanname_no': beyanname_no,
                    'Seçim_Nedenleri': ', '.join(self.selection_reasons.get(beyanname_no, []))
                }
                
                # Tarih bilgisi varsa ekle
                if date_column and date_column in matching_rows.columns:
                    summary_row['Tarih'] = matching_rows.iloc[0][date_column]
                
                summary_data.append(summary_row)
            
            # Excel Writer oluştur ve chunk boyutunu arttır
            from openpyxl.utils.exceptions import IllegalCharacterError
            
            # Excel engine options
            engine_kwargs = {
                'options': {'strings_to_urls': False}  # Güvenlik için URL dönüşümünü devre dışı bırak
            }
            
            # Excel Writer'ı oluştur
            writer = pd.ExcelWriter(output_path, engine='openpyxl', engine_kwargs=engine_kwargs)
            
            # Özet bilgileri Excel'e yaz
            try:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Örnekleme Özeti', index=False)
            except Exception as e:
                print(f"Özet sayfası yazma hatası: {str(e)}")
                # Hata durumunda basitleştirilmiş veri oluştur
                pd.DataFrame({'Beyanname_no': list(self.selected_beyannames)}).to_excel(
                    writer, sheet_name='Örnekleme Özeti', index=False)
            
            # Tam beyanname detaylarını Excel'e yaz
            try:
                available_columns = [col for col in needed_columns if col in selected_unique_df.columns]
                if available_columns:
                    # İllegal karakterleri temizleme
                    for col in available_columns:
                        if selected_unique_df[col].dtype == object:
                            selected_unique_df[col] = selected_unique_df[col].astype(str).apply(
                                lambda x: ''.join([c for c in x if ord(c) < 65535]))
                    
                    # Büyük veri için batching uygula (bellek optimizasyonu)
                    if len(selected_unique_df) > 5000:
                        # Veriyi parçalara ayırarak yaz
                        chunk_size = 5000
                        for i in range(0, len(selected_unique_df), chunk_size):
                            chunk = selected_unique_df.iloc[i:i+chunk_size]
                            if i == 0:
                                # İlk chunk için yeni sayfa oluştur
                                chunk[available_columns].to_excel(writer, sheet_name='Beyanname Detayları', index=False)
                            else:
                                # Diğer chunk'lar için ayrı sayfalar oluştur
                                chunk[available_columns].to_excel(writer, sheet_name=f'Beyanname Detayları_{i//chunk_size+1}', index=False)
                    else:
                        # Normal boyuttaki veri için standart yazma
                        selected_unique_df[available_columns].to_excel(writer, sheet_name='Beyanname Detayları', index=False)
                else:
                    # Sütun bulunamazsa en azından beyanname numaralarını yaz
                    pd.DataFrame({'Beyanname_no': list(self.selected_beyannames)}).to_excel(
                        writer, sheet_name='Beyanname Detayları', index=False)
            except Exception as e:
                print(f"Detay sayfası yazma hatası: {str(e)}")
                # Hata durumunda basitleştirilmiş veri oluştur
                pd.DataFrame({'Beyanname_no': list(self.selected_beyannames)}).to_excel(
                    writer, sheet_name='Beyanname Detayları', index=False)
            
            # İstatistik bilgilerini Excel'e yaz
            try:
                stats_df = pd.DataFrame({
                    'İstatistik': ['Toplam Beyanname Sayısı', 'Seçilen Beyanname Sayısı', 'Hedef Örnekleme Sayısı', 'Seçim Oranı (%)'],
                    'Değer': [
                        self.sampling_stats.get('total_beyannames', 0),
                        len(self.selected_beyannames),
                        self.sampling_stats.get('target_sample_count', 0),
                        round(len(self.selected_beyannames) / max(self.sampling_stats.get('total_beyannames', 1), 1) * 100, 2)
                    ]
                })
                
                stats_df.to_excel(writer, sheet_name='İstatistikler', index=False)
            except Exception as e:
                print(f"İstatistik sayfası yazma hatası: {str(e)}")
                # Basit istatistik sayfası oluştur
                pd.DataFrame({'İstatistik': ['Seçilen Beyanname Sayısı'], 
                             'Değer': [len(self.selected_beyannames)]}).to_excel(
                                 writer, sheet_name='İstatistikler', index=False)
            
            # Excel'i kaydet
            writer.close()
            writer = None
            
            # Belleği temizle
            import gc
            del selected_unique_df
            del summary_data
            gc.collect()
            
            return output_path
            
        except Exception as e:
            import traceback
            print(f"Excel aktarma hatası: {str(e)}")
            print(traceback.format_exc())
            
            # Writer nesnesi açık kaldıysa kapat
            if writer is not None:
                try:
                    writer.close()
                except:
                    pass
            
            raise ValueError(f"Excel oluşturulurken hata: {str(e)}")
        
        finally:
            # Writer açık kalmış olabilir, temizlik yap
            if writer is not None:
                try:
                    writer.close()
                except:
                    pass
            
            # Belleği temizle
            import gc
            gc.collect()
    
    def format_excel_report(self, output_path):
        """
        Excel raporunu biçimlendirir
        
        Args:
            output_path (str): Excel dosyasının yolu
        """
        # Zaman aşımı önlemi
        import signal
        
        # Timeout handler
        def timeout_handler(signum, frame):
            raise TimeoutError("Excel biçimlendirme işlemi zaman aşımına uğradı")
        
        # 30 saniye timeout ayarla
        old_handler = signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(30)  # 30 saniye
        
        wb = None
        
        try:
            # Excel dosyasını yükle
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            
            # Özet sayfasını biçimlendir
            if 'Örnekleme Özeti' in wb.sheetnames:
                ws = wb['Örnekleme Özeti']
                
                if ws.max_row > 0:  # Sayfada veri olduğunu kontrol et
                    # Başlıkları biçimlendir
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Tüm hücreler için temel stiller
                    default_font = Font(name='Calibri', size=11)
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Sütun genişliklerini ayarla
                    ws.column_dimensions['A'].width = 20  # Beyanname No
                    if 'B' in ws.column_dimensions:
                        ws.column_dimensions['B'].width = 15  # Tarih
                    if 'C' in ws.column_dimensions:
                        ws.column_dimensions['C'].width = 60  # Seçim Nedenleri
                    
                    # Tüm hücreler için kenarlık ayarla
                    thin_border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin'))
                    
                    # Satır sayısını kontrol et ve sınırla
                    max_rows_to_format = min(ws.max_row, 1000)  # En fazla 1000 satır biçimlendir
                    
                    for row in ws.iter_rows(min_row=1, max_row=max_rows_to_format, min_col=1, max_col=min(3, ws.max_column)):
                        for cell in row:
                            cell.border = thin_border
                            if row[0].row > 1:  # Başlık satırı değilse
                                cell.font = default_font
                    
                    # Alternatif satır renklendirme
                    alternate_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
                    for row_idx in range(2, max_rows_to_format + 1):
                        if row_idx % 2 == 0:  # Çift satırlar
                            for col_idx in range(1, min(4, ws.max_column + 1)):
                                if col_idx <= ws.max_column:  # Sütun sınırlarını kontrol et
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    cell.fill = alternate_fill
            
            # Detay sayfasını biçimlendir
            for sheet_name in wb.sheetnames:
                if 'Beyanname Detayları' in sheet_name:
                    ws = wb[sheet_name]
                    
                    if ws.max_row > 0:  # Sayfada veri olduğunu kontrol et
                        # Başlıkları biçimlendir
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        for cell in ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Sütun genişliklerini otomatik ayarla (ilk 10 sütun için)
                        for col_idx in range(1, min(11, ws.max_column + 1)):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                            ws.column_dimensions[col_letter].width = 15
                        
                        # Tüm hücreler için kenarlık ayarla
                        thin_border = Border(left=Side(style='thin'), 
                                            right=Side(style='thin'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))
                        
                        # Satır sayısını kontrol et ve sınırla
                        max_rows_to_format = min(ws.max_row, 1000)  # En fazla 1000 satır biçimlendir
                        max_cols_to_format = min(ws.max_column, 20)  # En fazla 20 sütun biçimlendir
                        
                        for row in ws.iter_rows(min_row=1, max_row=max_rows_to_format, min_col=1, max_col=max_cols_to_format):
                            for cell in row:
                                cell.border = thin_border
                        
                        # Alternatif satır renklendirme
                        alternate_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
                        for row_idx in range(2, max_rows_to_format + 1):
                            if row_idx % 2 == 0:  # Çift satırlar
                                for col_idx in range(1, max_cols_to_format + 1):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    cell.fill = alternate_fill
            
            # İstatistik sayfasını biçimlendir
            if 'İstatistikler' in wb.sheetnames:
                ws = wb['İstatistikler']
                
                if ws.max_row > 0:  # Sayfada veri olduğunu kontrol et
                    # Başlıkları biçimlendir
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    header_alignment = Alignment(horizontal='center', vertical='center')
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Sütun genişliklerini ayarla
                    ws.column_dimensions['A'].width = 25  # İstatistik
                    if 'B' in ws.column_dimensions:
                        ws.column_dimensions['B'].width = 15  # Değer
                    
                    # Tüm hücreler için kenarlık ve hizalama ayarla
                    thin_border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin'))
                    
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(2, ws.max_column)):
                        for cell in row:
                            cell.border = thin_border
                            if cell.column == 2:  # Değer sütunu
                                cell.alignment = Alignment(horizontal='center')
            
            # Dosyayı kaydet
            wb.save(output_path)
            wb = None  # Nesneyi serbest bırak
            
            # Belleği temizle
            import gc
            gc.collect()
            
        except TimeoutError as e:
            print(f"Excel biçimlendirme zaman aşımı: {str(e)}")
            # Zaman aşımı durumunda biçimlendirmeyi atla ama Excel dosyası hala kullanılabilir
            
        except Exception as e:
            import traceback
            print(f"Excel biçimlendirme hatası: {str(e)}")
            print(traceback.format_exc())
            # Hata olsa bile rapor oluşturuldu, sadece biçimlendirme yapılamadı
            
        finally:
            # Timeout'u sıfırla
            signal.alarm(0)
            signal.signal(signal.SIGALRM, old_handler)
            
            # Workbook nesnesi açık kaldıysa kapat
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
            
            # Belleği temizle
            import gc
            gc.collect() 